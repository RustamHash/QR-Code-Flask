"""
Сервис для обработки Excel файлов.
"""

import io
from typing import List, Tuple, Union
import pandas as pd
import logging
import re
from openpyxl import load_workbook

from .exceptions import ExcelProcessingError

logger = logging.getLogger(__name__)


def convert_scientific_notation(value: str) -> str:
    """
    Преобразует число из экспоненциальной записи в обычный формат.
    
    Args:
        value: Строка, которая может содержать число в экспоненциальной записи
        
    Returns:
        str: Число в обычном формате без экспоненциальной записи
    """
    if not isinstance(value, str):
        value = str(value)
    
    # Нормализуем: заменяем запятую на точку для десятичного разделителя
    value_normalized = value.replace(',', '.')
    
    # Проверяем, содержит ли строка экспоненциальную запись
    # Поддерживаем форматы: 1.02206e+15, 1.02206E+15, 1.02206e15, 1.02206E-15 и т.д.
    # Паттерн: опциональный знак, цифры, опциональная точка и цифры, e/E, опциональный знак, цифры
    pattern = r'^([+-]?\d+\.?\d*)[eE]([+-]?\d+)$'
    match = re.match(pattern, value_normalized.strip())
    
    if match:
        try:
            base = float(match.group(1))
            exponent = int(match.group(2))
            result = base * (10 ** exponent)
            
            # Если результат целое число, возвращаем без десятичной точки
            if result.is_integer():
                return str(int(result))
            else:
                # Для дробных чисел используем форматирование, чтобы избежать экспоненциальной записи
                # Определяем количество значащих цифр после запятой
                formatted = f"{result:.15f}".rstrip('0').rstrip('.')
                # Если все еще в экспоненциальной записи, используем другой подход
                if 'e' in formatted.lower():
                    # Для очень маленьких чисел оставляем как есть или используем форматирование
                    formatted = f"{result:.15f}".rstrip('0').rstrip('.')
                return formatted
        except (ValueError, OverflowError):
            # Если не удалось преобразовать, возвращаем исходное значение
            return value
    
    return value


def _read_excel_with_precision(excel_file: Union[io.BytesIO, str], column_index: int = 0) -> List[str]:
    """
    Читает данные из Excel файла с сохранением точности чисел через openpyxl.
    
    Args:
        excel_file: Путь к файлу или BytesIO объект
        column_index: Индекс колонки для чтения
        
    Returns:
        List[str]: Список данных из колонки
    """
    try:
        # Используем openpyxl для чтения точных значений
        if isinstance(excel_file, io.BytesIO):
            excel_file.seek(0)
            wb = load_workbook(excel_file, data_only=False)
        else:
            wb = load_workbook(excel_file, data_only=False)
        
        ws = wb.active
        data_list = []
        
        for row in ws.iter_rows(min_row=1, values_only=False):
            if column_index >= len(row):
                continue
            cell = row[column_index]
            if cell.value is not None:
                # Получаем точное значение из ячейки (openpyxl возвращает точное значение)
                value = cell.value
                # Преобразуем в строку, сохраняя точность
                if isinstance(value, int):
                    # Для целых чисел просто преобразуем в строку
                    data_list.append(str(value))
                elif isinstance(value, float):
                    # Для float проверяем, является ли оно целым числом
                    if value.is_integer():
                        # Преобразуем в int для точности, затем в строку
                        data_list.append(str(int(value)))
                    else:
                        # Для дробных чисел используем обычное преобразование
                        data_list.append(str(value))
                else:
                    # Для других типов (строки, даты и т.д.)
                    data_list.append(str(value))
        
        wb.close()
        return data_list
    except Exception as e:
        logger.warning(f"Не удалось прочитать через openpyxl, используем pandas: {e}")
        return None


def read_data_from_excel(excel_file: Union[io.BytesIO, str], column_index: int = 0) -> List[str]:
    """
    Читает данные из Excel файла.

    Args:
        excel_file: Путь к файлу или BytesIO объект
        column_index: Индекс колонки для чтения (по умолчанию 0 - первая колонка)

    Returns:
        List[str]: Список данных из колонки

    Raises:
        ExcelProcessingError: если не удалось прочитать файл
    """
    try:
        # Сначала пытаемся прочитать через openpyxl для точности
        data_list = _read_excel_with_precision(excel_file, column_index)
        
        if data_list is None:
            # Если не получилось, используем pandas
            df = pd.read_excel(excel_file, sheet_name=0, header=None, dtype=str, keep_default_na=False)

            if df.empty:
                raise ExcelProcessingError("Excel файл пуст")

            if column_index >= len(df.columns):
                raise ExcelProcessingError(
                    f"Колонка с индексом {column_index} не существует. "
                    f"Доступно колонок: {len(df.columns)}"
                )

            column_data = df.iloc[:, column_index]
            data_list = [convert_scientific_notation(str(item)).strip() for item in column_data if item and str(item).strip()]

        if not data_list:
            raise ExcelProcessingError("Не найдено данных в указанной колонке")

        # Применяем преобразование экспоненциальной записи только если данные были прочитаны через pandas
        # (данные из openpyxl уже точные и не требуют преобразования)
        if data_list and any('e' in str(item).lower() or 'E' in str(item) for item in data_list):
            data_list = [convert_scientific_notation(item).strip() for item in data_list if item and item.strip()]
        else:
            # Просто убираем пробелы
            data_list = [str(item).strip() for item in data_list if item and str(item).strip()]

        logger.info(f"Прочитано {len(data_list)} записей из Excel файла")
        return data_list

    except pd.errors.EmptyDataError as e:
        logger.error(f"Excel файл пуст: {e}")
        raise ExcelProcessingError("Excel файл пуст или поврежден") from e
    except Exception as e:
        error_msg = str(e).lower()
        if "excel" in error_msg or "xlsx" in error_msg or "xls" in error_msg:
            logger.error(f"Ошибка при чтении Excel файла: {e}")
            raise ExcelProcessingError(f"Не удалось прочитать Excel файл: {e}") from e
        logger.error(f"Неожиданная ошибка при обработке Excel файла: {e}", exc_info=True)
        raise ExcelProcessingError(f"Ошибка при обработке Excel файла: {e}") from e


def _read_pairs_with_precision(excel_file: Union[io.BytesIO, str]) -> List[Tuple[str, str]]:
    """
    Читает пары ключ-значение из Excel с сохранением точности через openpyxl.
    
    Args:
        excel_file: Путь к файлу или BytesIO объект
        
    Returns:
        List[Tuple[str, str]]: Список пар или None в случае ошибки
    """
    try:
        if isinstance(excel_file, io.BytesIO):
            excel_file.seek(0)
            wb = load_workbook(excel_file, data_only=False)
        else:
            wb = load_workbook(excel_file, data_only=False)
        
        ws = wb.active
        pairs = []
        
        for row in ws.iter_rows(min_row=1, values_only=False):
            if len(row) < 2:
                continue
                
            cell_a = row[0]
            cell_b = row[1]
            
            if cell_a.value is None or cell_b.value is None:
                continue
            
            # Преобразуем значения с сохранением точности
            def format_value(value):
                if isinstance(value, int):
                    # Для целых чисел просто преобразуем в строку
                    return str(value)
                elif isinstance(value, float):
                    # Для float проверяем, является ли оно целым числом
                    if value.is_integer():
                        # Преобразуем в int для точности, затем в строку
                        return str(int(value))
                    else:
                        return str(value)
                else:
                    return str(value)
            
            a_val = format_value(cell_a.value).strip()
            b_val = format_value(cell_b.value).strip()
            
            if a_val and b_val:
                pairs.append((a_val, b_val))
        
        wb.close()
        return pairs
    except Exception as e:
        logger.warning(f"Не удалось прочитать через openpyxl, используем pandas: {e}")
        return None


def read_key_value_pairs_from_excel(excel_file: Union[io.BytesIO, str]) -> List[Tuple[str, str]]:
    """
    Читает пары (колонка A, колонка B) из Excel файла строго из двух колонок.

    Args:
        excel_file: Путь к файлу или BytesIO объект

    Returns:
        List[Tuple[str, str]]: Список кортежей (значение_из_A, значение_из_B)

    Raises:
        ExcelProcessingError: если не удалось прочитать файл или формат не соответствует
    """
    try:
        # Сначала пытаемся прочитать через openpyxl для точности
        pairs = _read_pairs_with_precision(excel_file)
        
        if pairs is None:
            # Если не получилось, используем pandas
            df = pd.read_excel(excel_file, sheet_name=0, header=None, dtype=str, keep_default_na=False)

            if df.empty:
                raise ExcelProcessingError("Excel файл пуст")

            # Убираем полностью пустые колонки
            df = df.loc[:, (df != '').any(axis=0)]

            if len(df.columns) != 2:
                raise ExcelProcessingError(
                    f"Требуется строго 2 колонки (A и B). Найдено колонок: {len(df.columns)}"
                )

            col_a = df.iloc[:, 0]
            col_b = df.iloc[:, 1]

            pairs = []
            for idx in range(len(df)):
                a_val_raw = str(col_a.iloc[idx]).strip() if col_a.iloc[idx] and str(col_a.iloc[idx]).strip() else ""
                b_val_raw = str(col_b.iloc[idx]).strip() if col_b.iloc[idx] and str(col_b.iloc[idx]).strip() else ""
                
                a_val = convert_scientific_notation(a_val_raw).strip() if a_val_raw else ""
                b_val = convert_scientific_notation(b_val_raw).strip() if b_val_raw else ""

                if a_val and b_val:
                    pairs.append((a_val, b_val))
        
        # Применяем преобразование экспоненциальной записи на всякий случай
        pairs = [(convert_scientific_notation(a).strip(), convert_scientific_notation(b).strip()) 
                 for a, b in pairs if a.strip() and b.strip()]

        if not pairs:
            raise ExcelProcessingError("Не найдено пар значений: заполните обе колонки A и B")

        logger.info(f"Прочитано {len(pairs)} пар значений из Excel файла")
        return pairs

    except pd.errors.EmptyDataError as e:
        logger.error(f"Excel файл пуст: {e}")
        raise ExcelProcessingError("Excel файл пуст или поврежден") from e
    except ExcelProcessingError:
        raise
    except Exception as e:
        error_msg = str(e).lower()
        if "excel" in error_msg or "xlsx" in error_msg or "xls" in error_msg:
            logger.error(f"Ошибка при чтении Excel файла: {e}")
            raise ExcelProcessingError(f"Не удалось прочитать Excel файл: {e}") from e
        logger.error(f"Неожиданная ошибка при обработке Excel файла: {e}", exc_info=True)
        raise ExcelProcessingError(f"Ошибка при обработке Excel файла: {e}") from e

